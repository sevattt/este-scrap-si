"""
Exporta resultados a Excel profesional con graficas, colores y formato empresarial.
"""
import os
import re
import json
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1


def export_data(df, raw_data, output_dir, base_name, fmt):
    os.makedirs(output_dir, exist_ok=True)
    paths = []
    if fmt in ("excel", "all"):
        paths.append(_excel(df, output_dir, base_name))
    if fmt in ("csv", "all"):
        paths.append(_csv(df, output_dir, base_name))
    if fmt in ("json", "all"):
        paths.append(_json(df, output_dir, base_name))
    return paths


def _excel(df, output_dir, base_name):
    path = os.path.join(output_dir, f"{base_name}.xlsx")
    writer = pd.ExcelWriter(path, engine="openpyxl")

    productos_df = _build_productos(df)

    # Hoja portada
    pd.DataFrame().to_excel(writer, sheet_name="Portada", index=False)

    if not productos_df.empty:
        productos_df.to_excel(writer, sheet_name="Productos", index=False)

    _summary(df, productos_df).to_excel(writer, sheet_name="Resumen", index=False)
    df.to_excel(writer, sheet_name="Datos completos", index=False)

    for tipo in df["tipo"].unique():
        sub = df[df["tipo"] == tipo].reset_index(drop=True)
        sub.to_excel(writer, sheet_name=tipo[:31], index=False)

    writer.close()
    _style_workbook(path, productos_df)
    return path


def _parse_precio(texto):
    t = str(texto).strip()
    tiene_simbolo = any(s in t for s in ["$", "€", "£", "¥", "USD", "EUR", "COP", "MXN"])
    if not tiene_simbolo:
        return None
    solo_nums = re.sub(r"[^\d.,]", "", t)
    if not solo_nums:
        return None
    try:
        if re.search(r",\d{2}$", solo_nums):
            solo_nums = solo_nums.replace(".", "").replace(",", ".")
        elif re.search(r"\.\d{2}$", solo_nums):
            solo_nums = solo_nums.replace(",", "")
        else:
            solo_nums = solo_nums.replace(",", "").replace(".", "")
        val = float(solo_nums)
        if val <= 0 or val > 500000:
            return None
        return val
    except Exception:
        return None


def _build_productos(df):
    titulos = df[df["tipo"] == "titulo"]["dato"].tolist()
    precios_raw = df[df["tipo"] == "precio"]["dato"].tolist()

    precios_limpios = []
    for p in precios_raw:
        val = _parse_precio(p)
        if val is not None:
            precios_limpios.append({"precio_texto": p.strip(), "precio_num": val})

    links = df[df["tipo"] == "link"]["dato"].tolist()
    links_prod = [l for l in links if any(x in l for x in ["/dp/", "/product", "/item", "p="])]
    if not links_prod:
        links_prod = links

    if not titulos and not precios_limpios:
        return pd.DataFrame()

    rows = []
    max_len = max(len(titulos), len(precios_limpios))
    for i in range(min(max_len, 500)):
        titulo = titulos[i] if i < len(titulos) else "—"
        precio_info = precios_limpios[i] if i < len(precios_limpios) else None
        link = links_prod[i] if i < len(links_prod) else "—"
        rows.append({
            "N°":           i + 1,
            "Producto":     titulo[:120],
            "Precio":       precio_info["precio_texto"] if precio_info else "—",
            "Precio (USD)": precio_info["precio_num"] if precio_info else None,
            "Link":         link,
            "Fuente":       df["fuente"].iloc[0] if not df.empty else "—",
            "Fecha":        datetime.now().strftime("%Y-%m-%d"),
        })

    productos_df = pd.DataFrame(rows)
    productos_df = productos_df.sort_values("Precio (USD)", na_position="last").reset_index(drop=True)
    productos_df["N°"] = range(1, len(productos_df) + 1)
    return productos_df


def _summary(df, productos_df):
    rows = [
        {"Métrica": "REPORTE DE SCRAPING", "Valor": ""},
        {"Métrica": "Generado por", "Valor": "OctoScraper Bot"},
        {"Métrica": "Fecha", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Métrica": "Fuente", "Valor": df["fuente"].iloc[0] if not df.empty else "—"},
        {"Métrica": "Total registros", "Valor": len(df)},
        {"Métrica": "Páginas scrapeadas", "Valor": df["url_pagina"].nunique()},
        {"Métrica": "", "Valor": ""},
    ]

    if not productos_df.empty and "Precio (USD)" in productos_df.columns:
        nums = productos_df["Precio (USD)"].dropna()
        if not nums.empty:
            rows += [
                {"Métrica": "ANÁLISIS DE PRECIOS", "Valor": ""},
                {"Métrica": "Precio más bajo",  "Valor": f"${nums.min():,.2f}"},
                {"Métrica": "Precio más alto",  "Valor": f"${nums.max():,.2f}"},
                {"Métrica": "Precio promedio",  "Valor": f"${nums.mean():,.2f}"},
                {"Métrica": "Precio mediana",   "Valor": f"${nums.median():,.2f}"},
                {"Métrica": "Total productos",  "Valor": len(nums)},
                {"Métrica": "", "Valor": ""},
                {"Métrica": "TOP 3 MÁS BARATOS", "Valor": "Precio"},
            ]
            for _, row in productos_df.dropna(subset=["Precio (USD)"]).head(3).iterrows():
                rows.append({"Métrica": f"  {row['Producto'][:60]}", "Valor": row["Precio"]})
            rows += [
                {"Métrica": "", "Valor": ""},
                {"Métrica": "TOP 3 MÁS CAROS", "Valor": "Precio"},
            ]
            for _, row in productos_df.dropna(subset=["Precio (USD)"]).tail(3).iterrows():
                rows.append({"Métrica": f"  {row['Producto'][:60]}", "Valor": row["Precio"]})

    rows += [
        {"Métrica": "", "Valor": ""},
        {"Métrica": "DESGLOSE POR TIPO", "Valor": "Cantidad"},
    ]
    for tipo, n in df["tipo"].value_counts().items():
        rows.append({"Métrica": f"  {tipo}", "Valor": n})

    return pd.DataFrame(rows)


def _style_workbook(path, productos_df):
    wb = load_workbook(path)

    # Colores corporativos
    COLOR_DARK    = "0a0a1a"
    COLOR_GREEN   = "00cc66"
    COLOR_GREEN2  = "00ff88"
    COLOR_ACCENT  = "6366f1"
    COLOR_WHITE   = "FFFFFF"
    COLOR_LIGHT   = "f0fff4"
    COLOR_YELLOW  = "fff3cd"
    COLOR_RED_L   = "fde8e8"
    COLOR_GREEN_L = "d4edda"

    def make_border(color="DDDDDD"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def make_fill(color):
        return PatternFill("solid", start_color=color, fgColor=color)

    for ws in wb.worksheets:
        name = ws.title

        if name == "Portada":
            _create_portada(ws, COLOR_DARK, COLOR_GREEN2, COLOR_WHITE)
            continue

        is_prod    = name == "Productos"
        is_resumen = name == "Resumen"

        # Header
        hdr_color = COLOR_DARK if not is_prod else "0d4f2e"
        hdr_font_color = COLOR_GREEN2 if not is_prod else COLOR_WHITE

        for cell in ws[1]:
            cell.font      = Font(name="Calibri", bold=True, color=hdr_font_color, size=11)
            cell.fill      = make_fill(hdr_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = make_border("444444")

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # Filas de datos
        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                cell.border    = make_border()

                # Color alternado
                if i % 2 == 0:
                    cell.fill = make_fill(COLOR_LIGHT if is_prod else "f8f9fa")

                # Colorear precios en hoja Productos
                if is_prod and cell.column == 4 and cell.value:
                    try:
                        val = float(cell.value)
                        nums = productos_df["Precio (USD)"].dropna()
                        if not nums.empty:
                            p33 = nums.quantile(0.33)
                            p66 = nums.quantile(0.66)
                            if val <= p33:
                                cell.fill = make_fill("c8f7dc")
                                cell.font = Font(name="Calibri", color="155724", bold=True)
                            elif val <= p66:
                                cell.fill = make_fill("fff3cd")
                                cell.font = Font(name="Calibri", color="856404", bold=True)
                            else:
                                cell.fill = make_fill("f8d7da")
                                cell.font = Font(name="Calibri", color="842029", bold=True)
                    except Exception:
                        pass

                # Resumen — resaltar secciones
                if is_resumen and cell.column == 1 and cell.value:
                    v = str(cell.value)
                    if v in ("REPORTE DE SCRAPING", "ANÁLISIS DE PRECIOS",
                             "TOP 3 MÁS BARATOS", "TOP 3 MÁS CAROS", "DESGLOSE POR TIPO"):
                        cell.font = Font(name="Calibri", bold=True, color=COLOR_GREEN, size=11)
                        cell.fill = make_fill("0a1a0f")

        # Fila más barata resaltada en verde brillante
        if is_prod and ws.max_row > 2:
            for cell in ws[2]:
                cell.fill = make_fill("a8f0c6")
                cell.font = Font(name="Calibri", bold=True, color="0a3d1f", size=11)

        # Ancho columnas
        for col in ws.columns:
            max_w = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_w + 3, 12), 55)

        # Agregar gráfica de precios en hoja Productos
        if is_prod and not productos_df.empty and ws.max_row > 3:
            _add_price_chart(ws, productos_df)

    wb.save(path)


def _create_portada(ws, bg_color, accent_color, white):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30

    # Fondo oscuro
    for row in range(1, 40):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", start_color=bg_color, fgColor=bg_color)

    # Título principal
    ws.merge_cells("B3:H5")
    title = ws["B3"]
    title.value = "🕷️ OCTOSCRAPER BOT"
    title.font = Font(name="Calibri", bold=True, size=32, color=accent_color)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Subtítulo
    ws.merge_cells("B6:H7")
    sub = ws["B6"]
    sub.value = "Reporte de Extracción de Datos"
    sub.font = Font(name="Calibri", size=16, color="aaaaaa")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # Info
    ws.row_dimensions[9].height = 5
    info = [
        ("Generado:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Herramienta:", "OctoScraper Bot v1.0"),
        ("Motor:", "requests + BeautifulSoup"),
        ("Exportado:", "Excel + CSV + JSON"),
    ]
    for i, (label, value) in enumerate(info, 10):
        ws.row_dimensions[i].height = 22
        lc = ws.cell(row=i, column=2, value=label)
        vc = ws.cell(row=i, column=3, value=value)
        lc.font = Font(name="Calibri", bold=True, color="888888", size=11)
        vc.font = Font(name="Calibri", color=accent_color, size=11)

    # Footer
    ws.merge_cells("B35:H36")
    footer = ws["B35"]
    footer.value = "Los datos mostrados fueron extraídos automáticamente. Verifique la información antes de tomar decisiones."
    footer.font = Font(name="Calibri", size=9, color="555555", italic=True)
    footer.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_price_chart(ws, productos_df):
    try:
        top10 = productos_df.dropna(subset=["Precio (USD)"]).head(10)
        if len(top10) < 2:
            return

        chart = BarChart()
        chart.type = "col"
        chart.title = "Top 10 Productos — Comparación de Precios"
        chart.y_axis.title = "Precio (USD)"
        chart.x_axis.title = "Producto"
        chart.style = 10
        chart.height = 12
        chart.width  = 22

        # Los precios están en columna D (col 4), filas 2 a 11
        data_ref = Reference(ws, min_col=4, min_row=1, max_row=min(11, ws.max_row))
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=min(11, ws.max_row))

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "00cc66"

        # Insertar después de los datos
        ws.add_chart(chart, f"A{ws.max_row + 3}")
    except Exception:
        pass


def _csv(df, output_dir, base_name):
    path = os.path.join(output_dir, f"{base_name}.csv")
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def _json(df, output_dir, base_name):
    path = os.path.join(output_dir, f"{base_name}.json")
    out = {
        "meta": {
            "generated_at":  datetime.now().isoformat(),
            "tool":          "OctoScraper Bot v1.0",
            "total_records": len(df),
            "fuentes":       df["fuente"].unique().tolist(),
        },
        "por_tipo": {
            tipo: df[df["tipo"] == tipo][["fuente", "url_pagina", "dato", "atributo"]].to_dict("records")
            for tipo in df["tipo"].unique()
        },
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    return path
